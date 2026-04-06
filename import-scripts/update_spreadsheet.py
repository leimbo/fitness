#!/usr/bin/env python3
"""
Health dashboard spreadsheet updater.
Reads all imported JSON data and writes/updates health_dashboard.xlsx
with structured sheets for activities, daily health, training load,
and a Runna planned-vs-actual comparison.

Usage:
  python update_spreadsheet.py \
    --data-dir /path/to/data/ \
    --output /path/to/health_dashboard.xlsx
"""

import argparse
import json
import os
import sys
from datetime import datetime, date
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.series import SeriesLabel
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl --break-system-packages")
    sys.exit(1)


# ── Colours ──────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
ALT_FILL    = PatternFill("solid", fgColor="EEF2F7")
ACCENT      = PatternFill("solid", fgColor="2E75B6")
GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")

THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_json(path: str, default=None):
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return default if default is not None else {}


def style_header_row(ws, row: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def auto_width(ws, min_width=10, max_width=40):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_width, max(min_width, max_len + 2))


def pace_from_speed(speed_ms: float | None, unit: str = "km") -> str:
    """Convert m/s speed to pace string (mm:ss/km or /mile)."""
    if not speed_ms or speed_ms == 0:
        return ""
    if unit == "km":
        secs_per_km = 1000 / speed_ms
    else:
        secs_per_km = 1609.34 / speed_ms
    mins = int(secs_per_km // 60)
    secs = int(secs_per_km % 60)
    return f"{mins}:{secs:02d}"


def meters_to_km(m) -> float | None:
    if m is None:
        return None
    try:
        return round(float(m) / 1000, 2)
    except (TypeError, ValueError):
        return None


def secs_to_hms(seconds) -> str:
    if not seconds:
        return ""
    try:
        s = int(seconds)
        h, rem = divmod(s, 3600)
        m, sec = divmod(rem, 60)
        if h:
            return f"{h}:{m:02d}:{sec:02d}"
        return f"{m}:{sec:02d}"
    except (TypeError, ValueError):
        return ""


# ── Sheet builders ────────────────────────────────────────────────────────────

def build_activities_sheet(wb, garmin_dir: str, strava_dir: str):
    ws = wb.create_sheet("Activities")
    ws.freeze_panes = "A2"

    headers = [
        "Date", "Platform", "Name", "Type",
        "Distance (km)", "Duration", "Avg Pace (min/km)",
        "Avg HR (bpm)", "Max HR (bpm)", "Elev Gain (m)",
        "Calories", "Avg Power (W)", "Training Effect", "Activity ID"
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 30

    rows = []

    # --- Garmin activities ---
    garmin_acts = load_json(os.path.join(garmin_dir, "activities.json"), [])
    for act in garmin_acts:
        start_local = act.get("startTimeLocal", act.get("startTimeGMT", ""))
        act_date = start_local[:10] if start_local else ""
        rows.append({
            "date": act_date,
            "platform": "Garmin",
            "name": act.get("activityName", ""),
            "type": act.get("activityType", {}).get("typeKey", "") if isinstance(act.get("activityType"), dict) else act.get("activityType", ""),
            "distance_km": meters_to_km(act.get("distance")),
            "duration": secs_to_hms(act.get("duration")),
            "avg_pace": pace_from_speed(act.get("averageSpeed")),
            "avg_hr": act.get("averageHR"),
            "max_hr": act.get("maxHR"),
            "elev_gain": act.get("elevationGain"),
            "calories": act.get("calories"),
            "avg_power": act.get("avgPower"),
            "training_effect": act.get("aerobicTrainingEffect"),
            "activity_id": act.get("activityId"),
        })

    # --- Strava activities ---
    strava_acts = load_json(os.path.join(strava_dir, "activities_detailed.json"), [])
    for act in strava_acts:
        start = act.get("start_date_local", "")
        act_date = start[:10] if start else ""
        rows.append({
            "date": act_date,
            "platform": "Strava",
            "name": act.get("name", ""),
            "type": act.get("type", act.get("sport_type", "")),
            "distance_km": meters_to_km(act.get("distance")),
            "duration": secs_to_hms(act.get("moving_time")),
            "avg_pace": pace_from_speed(act.get("average_speed")),
            "avg_hr": act.get("average_heartrate"),
            "max_hr": act.get("max_heartrate"),
            "elev_gain": act.get("total_elevation_gain"),
            "calories": act.get("calories"),
            "avg_power": act.get("average_watts"),
            "training_effect": None,
            "activity_id": act.get("id"),
        })

    # Sort by date descending
    rows.sort(key=lambda r: r.get("date", ""), reverse=True)

    for i, r in enumerate(rows, start=2):
        row_data = [
            r["date"], r["platform"], r["name"], r["type"],
            r["distance_km"], r["duration"], r["avg_pace"],
            r["avg_hr"], r["max_hr"], r["elev_gain"],
            r["calories"], r["avg_power"], r["training_effect"], r["activity_id"]
        ]
        ws.append(row_data)
        if i % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).fill = ALT_FILL
        for col in range(1, len(headers) + 1):
            ws.cell(row=i, column=col).border = BORDER

    auto_width(ws)
    return len(rows)


def build_daily_health_sheet(wb, garmin_dir: str):
    ws = wb.create_sheet("Daily Health")
    ws.freeze_panes = "A2"

    headers = [
        "Date", "Resting HR", "HRV (ms)", "Body Battery (max)",
        "Avg Stress", "Sleep Duration (h)", "Sleep Score",
        "Deep Sleep (h)", "REM Sleep (h)", "Steps", "VO2max"
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 30

    daily = load_json(os.path.join(garmin_dir, "daily_metrics.json"), {})

    rows = []
    for day_str, data in sorted(daily.items(), reverse=True):
        # Sleep
        sleep = data.get("sleep", {})
        sleep_duration = None
        sleep_score = None
        deep_sleep = None
        rem_sleep = None
        if isinstance(sleep, dict):
            sleep_duration_s = sleep.get("sleepTimeSeconds")
            sleep_duration = round(sleep_duration_s / 3600, 2) if sleep_duration_s else None
            sleep_score = sleep.get("sleepScores", {}).get("overall", {}).get("value") if isinstance(sleep.get("sleepScores"), dict) else None
            deep_sleep_s = sleep.get("deepSleepSeconds")
            deep_sleep = round(deep_sleep_s / 3600, 2) if deep_sleep_s else None
            rem_sleep_s = sleep.get("remSleepSeconds")
            rem_sleep = round(rem_sleep_s / 3600, 2) if rem_sleep_s else None

        # Body battery
        bb = data.get("body_battery", {})
        bb_max = None
        if isinstance(bb, list) and bb:
            vals = [x.get("bodyBatteryLevel") for x in bb if x.get("bodyBatteryLevel") is not None]
            bb_max = max(vals) if vals else None
        elif isinstance(bb, dict):
            bb_max = bb.get("startTimestampGMT")  # fallback

        # HRV
        hrv = data.get("hrv", {})
        hrv_val = None
        if isinstance(hrv, dict):
            summary = hrv.get("hrvSummary", {})
            hrv_val = summary.get("lastNight") if isinstance(summary, dict) else None

        # Steps
        steps_data = data.get("steps", {})
        total_steps = None
        if isinstance(steps_data, list):
            total_steps = sum(s.get("steps", 0) for s in steps_data if isinstance(s, dict))
        elif isinstance(steps_data, dict):
            total_steps = steps_data.get("totalSteps")

        rows.append([
            day_str,
            data.get("resting_hr"),
            hrv_val,
            bb_max,
            data.get("avg_stress"),
            sleep_duration,
            sleep_score,
            deep_sleep,
            rem_sleep,
            total_steps,
            data.get("vo2max"),
        ])

    for i, row in enumerate(rows, start=2):
        ws.append(row)
        if i % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).fill = ALT_FILL
        # Colour-code sleep score
        sleep_score_cell = ws.cell(row=i, column=7)
        if sleep_score_cell.value is not None:
            try:
                score = float(sleep_score_cell.value)
                if score >= 80:
                    sleep_score_cell.fill = GREEN_FILL
                elif score < 60:
                    sleep_score_cell.fill = RED_FILL
                else:
                    sleep_score_cell.fill = YELLOW_FILL
            except (TypeError, ValueError):
                pass
        for col in range(1, len(headers) + 1):
            ws.cell(row=i, column=col).border = BORDER

    auto_width(ws)
    return len(rows)


def build_planned_vs_actual_sheet(wb, runna_dir: str, garmin_dir: str, strava_dir: str):
    ws = wb.create_sheet("Planned vs Actual")
    ws.freeze_panes = "A2"

    headers = [
        "Date", "Planned (Runna)", "Planned Distance", "Planned Duration",
        "Actual Activity", "Actual Distance (km)", "Actual Duration",
        "Actual Avg Pace", "Actual Avg HR", "Completed?"
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 30

    runna_by_date = load_json(os.path.join(runna_dir, "planned_workouts_by_date.json"), {})

    # Build actual activities lookup by date (prefer Garmin, fallback Strava)
    actual_by_date = {}
    strava_acts = load_json(os.path.join(strava_dir, "activities.json"), [])
    for act in strava_acts:
        d = act.get("start_date_local", "")[:10]
        if d:
            actual_by_date.setdefault(d, []).append({"source": "Strava", **act})

    garmin_acts = load_json(os.path.join(garmin_dir, "activities.json"), [])
    for act in garmin_acts:
        d = act.get("startTimeLocal", "")[:10]
        if d:
            actual_by_date.setdefault(d, []).append({"source": "Garmin", **act})

    rows = []
    all_dates = sorted(set(list(runna_by_date.keys()) + list(actual_by_date.keys())), reverse=True)

    for d in all_dates:
        planned = runna_by_date.get(d, {})
        actuals = actual_by_date.get(d, [])
        # Pick the most relevant actual (prefer runs)
        actual = next(
            (a for a in actuals if "run" in str(a.get("type", a.get("activityType", {}))).lower()),
            actuals[0] if actuals else {}
        )

        planned_dist = planned.get("planned_distance")
        planned_title = planned.get("title", planned.get("workout_type", ""))
        planned_dur = planned.get("planned_duration_min")

        if actual.get("source") == "Strava":
            actual_name = actual.get("name", "")
            actual_dist = meters_to_km(actual.get("distance"))
            actual_dur = secs_to_hms(actual.get("moving_time"))
            actual_pace = pace_from_speed(actual.get("average_speed"))
            actual_hr = actual.get("average_heartrate")
        elif actual.get("source") == "Garmin":
            actual_name = actual.get("activityName", "")
            actual_dist = meters_to_km(actual.get("distance"))
            actual_dur = secs_to_hms(actual.get("duration"))
            actual_pace = pace_from_speed(actual.get("averageSpeed"))
            actual_hr = actual.get("averageHR")
        else:
            actual_name = actual_dist = actual_dur = actual_pace = actual_hr = None

        completed = "✓" if actual else ("Rest" if planned.get("workout_type") == "rest" else "✗")

        rows.append([
            d, planned_title,
            f"{planned_dist} {planned.get('planned_distance_unit', '')}" if planned_dist else "",
            f"{planned_dur} min" if planned_dur else "",
            actual_name, actual_dist, actual_dur, actual_pace, actual_hr, completed
        ])

    for i, row in enumerate(rows, start=2):
        ws.append(row)
        completed_cell = ws.cell(row=i, column=10)
        if completed_cell.value == "✓":
            completed_cell.fill = GREEN_FILL
        elif completed_cell.value == "✗":
            completed_cell.fill = RED_FILL
        elif completed_cell.value == "Rest":
            completed_cell.fill = ALT_FILL
        if i % 2 == 0:
            for col in range(1, 10):
                if ws.cell(row=i, column=col).fill.fgColor.rgb == "00000000":
                    ws.cell(row=i, column=col).fill = ALT_FILL
        for col in range(1, len(headers) + 1):
            ws.cell(row=i, column=col).border = BORDER

    auto_width(ws)
    return len(rows)


def build_summary_sheet(wb, activity_count: int, day_count: int, last_import: dict):
    ws = wb.create_sheet("Summary", 0)

    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "Health & Fitness Dashboard"
    title_cell.font = Font(size=18, bold=True, color="1F3864")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws["A3"] = "Last updated"
    ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws["A4"] = "Activities imported"
    ws["B4"] = activity_count
    ws["A5"] = "Days of health data"
    ws["B5"] = day_count
    ws["A6"] = "Date range"
    ws["B6"] = f"{last_import.get('start', 'N/A')} → {last_import.get('end', 'N/A')}"

    ws["A8"] = "Platforms"
    for i, p in enumerate(last_import.get("platforms", []), start=9):
        ws[f"A{i}"] = f"  ✓ {p.title()}"

    for row in ws.iter_rows(min_row=3, max_row=15, min_col=1, max_col=2):
        for cell in row:
            cell.border = BORDER
            if cell.column == 1:
                cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Update health dashboard spreadsheet")
    parser.add_argument("--data-dir", required=True, help="Root data directory")
    parser.add_argument("--output", required=True, help="Output .xlsx file path")
    args = parser.parse_args()

    garmin_dir = os.path.join(args.data_dir, "garmin")
    strava_dir = os.path.join(args.data_dir, "strava")
    runna_dir  = os.path.join(args.data_dir, "runna")

    print(f"\n{'='*50}")
    print(f"Updating: {args.output}")
    print(f"{'='*50}\n")

    # Load last import metadata
    last_import = load_json(os.path.join(args.data_dir, "last_import.json"), {})

    wb = openpyxl.Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    print("Building Activities sheet...")
    activity_count = build_activities_sheet(wb, garmin_dir, strava_dir)
    print(f"  ✓ {activity_count} activities")

    print("Building Daily Health sheet...")
    day_count = build_daily_health_sheet(wb, garmin_dir)
    print(f"  ✓ {day_count} days")

    print("Building Planned vs Actual sheet...")
    build_planned_vs_actual_sheet(wb, runna_dir, garmin_dir, strava_dir)
    print("  ✓ Done")

    print("Building Summary sheet...")
    build_summary_sheet(wb, activity_count, day_count, last_import)
    print("  ✓ Done")

    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"\n✓ Saved → {args.output}")


if __name__ == "__main__":
    main()
